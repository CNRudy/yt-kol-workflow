#!/usr/bin/env python3
# -*- coding: utf-8 -*-
import sys, os, json
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
import openpyxl

root = os.path.dirname(os.path.abspath(__file__))
with open('nosource_cids.json') as f:
    recs = json.load(f)
cids = set(r['cid'] for r in recs)
print('无来源 cid:', len(cids))

# --- 1. 8/26 influencers_all.xlsx 来源关键词列 ---
p1 = os.path.join(root, 'output/20260826_153843_batch/influencers_all.xlsx')
wbk = openpyxl.load_workbook(p1, read_only=True, data_only=True)
ws = wbk.active
rows = ws.iter_rows(values_only=True)
header = next(rows)
print('\n[8/26 influencers_all] 表头含"来源":', [h for h in header if h and '来源' in str(h)])
ci_cid = next((i for i,h in enumerate(header) if h and 'channel id' in str(h).lower()), None)
ci_src = next((i for i,h in enumerate(header) if h and '来源关键词' in str(h)), None)
print('  Channel ID idx:', ci_cid, '| 来源关键词 idx:', ci_src)
m = {}
for row in rows:
    if ci_cid is None or ci_cid>=len(row): continue
    c=str(row[ci_cid] or '').strip()
    if c in cids:
        v=row[ci_src] if ci_src is not None and ci_src<len(row) else None
        m[c]=v
print(f'  在 8/26 influencers_all 匹配到: {len(m)} / {len(cids)}')
print('  来源关键词样例:', {c:str(v)[:40] for c,v in list(m.items())[:5]})
wbk.close()

# --- 2. us_baby_full_match 排序组列真实值 ---
p2 = os.path.join(root, 'output/us_baby_full_match/全表匹配_美国婴儿车载摄像头.xlsx')
wbk2 = openpyxl.load_workbook(p2, read_only=True, data_only=True)
ws2 = wbk2.active
rows2 = ws2.iter_rows(values_only=True)
header2 = next(rows2)
print('\n[us_baby_full_match] 排序组idx:', header2.index('排序组') if '排序组' in header2 else 'N/A')
ci_cid2 = header2.index('Channel ID')
seen=[]
cnt=0
for row in rows2:
    if ci_cid2>=len(row): continue
    c=str(row[ci_cid2] or '').strip()
    if c in cids:
        cnt+=1
        if len(seen)<8:
            seen.append(str(row[header2.index('排序组')]) if '排序组' in header2 else '?')
print(f'  匹配到: {cnt} / {len(cids)}')
print('  排序组值样例:', seen)
wbk2.close()
