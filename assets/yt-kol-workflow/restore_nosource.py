#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""把 273 条「来源关键词为空」记录按 8/26 婴儿车载摄像头关键词批次重新归属 WEMOH C1 匹配产品。"""
import os
import sys, os, json
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
import write_user_base as wb
import openpyxl
wb.LARK = '/Users/coscod/.workbuddy/binaries/node/cli-connector-packages/bin/lark-cli'
os.environ.setdefault('HTTPS_PROXY', 'http://127.0.0.1:7890')
os.environ.setdefault('HTTP_PROXY', 'http://127.0.0.1:7890')
os.environ['LARK_CLI_NO_PROXY_WARN'] = '1'

BASE_TOKEN = os.environ["FEISHU_BASE_TOKEN"]  # 从 .env / 环境变量读取，勿提交真实值
TABLE_ID  = 'tbl4rnFUM9jJXvCQ'
PROD_NAME = 'WEMOH C1 Baby Car Camera (婴儿车载摄像头, ASIN B0F6N9PHHW)'

def run(a): return wb.run(a)

def load_nosource():
    with open('nosource_cids.json') as f:
        return json.load(f)  # [{'rid','cid'}]

def load_src_from_826():
    """从 8/26 influencers_all.xlsx 取 cid -> 来源关键词(单字符串)"""
    p = os.path.join('output/20260826_153843_batch/influencers_all.xlsx')
    wbk = openpyxl.load_workbook(p, read_only=True, data_only=True)
    ws = wbk.active
    rows = ws.iter_rows(values_only=True)
    header = next(rows)
    ci_cid = header.index('Channel ID')
    ci_src = header.index('来源关键词')
    m = {}
    for row in rows:
        if ci_cid>=len(row): continue
        c=str(row[ci_cid] or '').strip()
        if c:
            v=row[ci_src]
            m[c]=str(v).strip() if v else ''
    wbk.close()
    return m

def load_match_from_full():
    """从 8 月全表匹配 Excel 取 cid -> 匹配字段"""
    p = os.path.join('output/us_baby_full_match/全表匹配_美国婴儿车载摄像头.xlsx')
    wbk = openpyxl.load_workbook(p, read_only=True, data_only=True)
    ws = wbk.active
    rows = ws.iter_rows(values_only=True)
    header = next(rows)
    idx = {k: header.index(k) for k in ['Channel ID','品牌匹配度','开发优先级','匹配关键词','内容契合类型','推荐理由','排序组']}
    m = {}
    for row in rows:
        c=str(row[idx['Channel ID']] or '').strip()
        if not c: continue
        def g(k):
            v=row[idx[k]]
            return str(v).strip() if v is not None else ''
        m[c] = {
            '品牌匹配度': g('品牌匹配度'),
            '开发优先级': g('开发优先级'),
            '匹配关键词': g('匹配关键词'),
            '内容契合类型': g('内容契合类型'),
            '推荐理由': g('推荐理由'),
            '匹配排序组': g('排序组'),
        }
    wbk.close()
    return m

def main():
    recs = load_nosource()
    print('无来源记录:', len(recs))
    cid2src = load_src_from_826()
    cid2match = load_match_from_full()
    cids = set(r['cid'] for r in recs)
    print('8/26 取到来源关键词:', sum(1 for c in cids if cid2src.get(c)))
    print('全表匹配取到字段:', sum(1 for c in cids if c in cid2match))

    # 1. 读字段现有选项
    fl = run(['base','+field-list','--as','user','--base-token',BASE_TOKEN,'--table-id',TABLE_ID,'--format','json'])
    src_field = next(f for f in fl['data']['fields'] if f['name']=='来源关键词')
    existing_opts = src_field.get('options', [])
    existing_names = [o['name'] for o in existing_opts]
    new_words = sorted({cid2src[c] for c in cids if cid2src.get(c) and cid2src[c] not in existing_names})
    print('来源关键词现有选项数:', len(existing_names))
    print('需新增来源关键词选项:', new_words)

    # 2. 补选项
    if new_words:
        new_opts = list(existing_opts) + [{'name': x} for x in new_words]
        upd = {'name': src_field['name'], 'type':'select',
               'multiple': src_field.get('multiple', False), 'options': new_opts}
        up = run(['base','+field-update','--as','user','--base-token',BASE_TOKEN,'--table-id',TABLE_ID,
                  '--field-id', src_field['id'], '--format','json','--yes','--json', json.dumps(upd)])
        print('field-update ok=', (up or {}).get('ok'), (up or {}).get('msg',''))

    # 3. 构造更新
    updates = {}
    for r in recs:
        cid = r['cid']; rid = r['rid']
        src = cid2src.get(cid, '')
        mt = cid2match.get(cid, {})
        def num(v):
            try: return int(float(v))
            except: return None
        updates[rid] = {
            '匹配产品': PROD_NAME,
            '来源关键词': src if src else None,
            '品牌匹配度': num(mt.get('品牌匹配度')),
            '开发优先级': mt.get('开发优先级') or None,
            '匹配关键词': mt.get('匹配关键词') or None,
            '内容契合类型': mt.get('内容契合类型') or None,
            '推荐理由': mt.get('推荐理由') or None,
            '匹配排序组': mt.get('匹配排序组') or None,
        }

    # 4. 先试 1 条
    first_rid = next(iter(updates))
    test = {first_rid: updates[first_rid]}
    jt = run(['base','+record-batch-update','--as','user','--base-token',BASE_TOKEN,'--table-id',TABLE_ID,
              '--format','json','--json', json.dumps({'update_records': test})])
    print('\n[试更新1条] ok=', (jt or {}).get('ok'), (jt or {}).get('msg',''), '| 字段数=', len(test[first_rid]))
    if not (jt and jt.get('ok')):
        print('试更新失败，停止'); return

    # 5. 全量
    ok=0; fail=0
    items=list(updates.items())
    for i in range(0, len(items), 200):
        chunk=dict(items[i:i+200])
        j = run(['base','+record-batch-update','--as','user','--base-token',BASE_TOKEN,'--table-id',TABLE_ID,
                 '--format','json','--json', json.dumps({'update_records': chunk})])
        if j and j.get('ok'):
            ok+=len(chunk); print(f'batch {i//200+1}: ok {len(chunk)}')
        else:
            fail+=len(chunk); print('FAIL:', json.dumps(j,ensure_ascii=False)[:400])
    print(f'\n=== 回填完成: 成功 {ok} / 失败 {fail} ===')

if __name__ == '__main__':
    main()
