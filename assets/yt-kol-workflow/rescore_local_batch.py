#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""离线用当前 active 产品画像，重算本地 batch 输出(influencers_all.xlsx)的匹配分。

适用场景：batch 跑了 --no-feishu，但 Phase D 用的是旧 active 画像（或你想换画像重算）。
不消耗任何 YouTube / 飞书配额，只重读本地 Excel + 重新打分。

用法:
    ./.venv/bin/python rescore_local_batch.py
    ./.venv/bin/python rescore_local_batch.py --batch-dir output/20260819_182645_batch
"""
import argparse
import json
import os
import sys

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

import openpyxl
from filter.scoring import get_active_profile, score_influencer

BATCH_DEFAULT = "output/20260819_182645_batch"

# 本地 influencers_all.xlsx 表头 → score_influencer 需要的 detail 键
DETAIL_MAP = {
    "亚马逊推广经验": "amazon_promo_level",
    "邮箱状态": "email_status",
    "代表视频互动率": "rep_video_engagement",
    "国家/地区": "country",
    "断更评估": "activity_status",
    "频道描述": "channel_description",
    "代表视频标题": "rep_video_title",
}


def read_xlsx(path):
    wb = openpyxl.load_workbook(path, read_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    if not rows:
        return [], []
    headers = list(rows[0])
    data = [dict(zip(headers, r)) for r in rows[1:]]
    return headers, data


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--batch-dir", default=BATCH_DEFAULT)
    args = ap.parse_args()

    base = args.batch_dir
    inf_path = os.path.join(base, "influencers_all.xlsx")
    vid_path = os.path.join(base, "influencer_videos_all.xlsx")
    if not os.path.exists(inf_path):
        print(f"✗ 找不到 {inf_path}")
        return

    profile = get_active_profile()
    if not profile:
        print("✗ 没有 active 产品画像，无法评分")
        return
    print(f"[画像] {profile.get('name')} | 市场={profile.get('markets')} | "
          f"min_views={profile.get('min_views')} min_eng={profile.get('min_engagement')}")

    inf_h, influencers = read_xlsx(inf_path)
    print(f"[读取] 网红 {len(influencers)} 个")

    vmap = {}
    if os.path.exists(vid_path):
        _, videos = read_xlsx(vid_path)
        for v in videos:
            cid = str(v.get("Channel ID", "") or "")
            if not cid:
                continue
            vmap.setdefault(cid, []).append({
                "title": v.get("Video Title", "") or "",
                "tags": v.get("Tags", "") or "",
                "description": v.get("字幕内容", "") or "",
                "view_count": v.get("Views", 0) or 0,
                "engagement_rate": v.get("互动率(%)", 0.0) or 0.0,
            })
        print(f"[读取] 视频 {len(videos)} 条，覆盖频道 {len(vmap)} 个")

    # 重算每个网红
    tier_counts = {}
    fit_sum = 0
    rescored = []
    for rec in influencers:
        cid = str(rec.get("Channel ID", "") or "")
        detail = {dk: rec.get(hk, "") for hk, dk in DETAIL_MAP.items()}
        # 数值字段兜底
        try:
            detail["rep_video_engagement"] = float(detail.get("rep_video_engagement") or 0)
        except (TypeError, ValueError):
            detail["rep_video_engagement"] = 0.0
        scored = score_influencer(detail, vmap.get(cid, []), profile)
        tier_counts[scored["dev_priority"]] = tier_counts.get(scored["dev_priority"], 0) + 1
        fit_sum += scored["brand_fit_score"]
        # 写回 rec 的 6 个评分列
        rec["匹配产品"] = scored["match_profile"]
        rec["品牌匹配度"] = scored["brand_fit_score"]
        rec["内容契合类型"] = scored["content_types"]
        rec["匹配关键词"] = scored["matched_keywords"]
        rec["开发优先级"] = scored["dev_priority"]
        rec["推荐理由"] = scored["recommend_reason"]
        rescored.append((cid, rec, scored))

    # 输出修正版 Excel
    out_path = os.path.join(base, "influencers_rescored.xlsx")
    new_wb = openpyxl.Workbook()
    ws = new_wb.active
    ws.append(list(inf_h))
    for rec in influencers:
        ws.append([rec.get(h, "") for h in inf_h])
    new_wb.save(out_path)

    avg_fit = round(fit_sum / max(1, len(influencers)), 1)
    print(f"\n✅ 重算完成，共 {len(influencers)} 个网红")
    print(f"   平均品牌匹配度: {avg_fit}")
    print("   分层分布: " + " / ".join(f"{t}档={tier_counts.get(t,0)}" for t in ("S","A","B","C")))
    print(f"   输出: {out_path}")

    # Top 15 by fit
    print("\n=== Top 15 按品牌匹配度 ===")
    top = sorted(rescored, key=lambda x: x[2]["brand_fit_score"], reverse=True)[:15]
    for cid, rec, sc in top:
        print(f"  {sc['brand_fit_score']:>3} {sc['dev_priority']} | {str(rec.get('Channel Name',''))[:32]:<32} | "
              f"国家={rec.get('国家/地区','?')} | 匹配词={sc['matched_keywords'][:50]}")

    # 国家分布
    cc = {}
    for rec in influencers:
        c = str(rec.get("国家/地区", "") or "?")
        cc[c] = cc.get(c, 0) + 1
    print("\n=== 国家分布 ===")
    for c, n in sorted(cc.items(), key=lambda x: -x[1]):
        print(f"  {c}: {n}")


if __name__ == "__main__":
    main()
