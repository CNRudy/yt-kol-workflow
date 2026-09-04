import os, requests, json

# 凭据从环境变量读取（.env 示例见 assets/yt-kol-workflow/.env.example），勿提交真实值
APP_ID = os.environ.get("FEISHU_APP_ID", "")
APP_SECRET = os.environ.get("FEISHU_APP_SECRET", "")
JB3 = os.environ.get("FEISHU_BASE_TOKEN", "")
TBL = "tbl1opmlkmE3veas"
XLSX = "/Users/coscod/WorkBuddy/coscod/YouTube 网红开发工作流系统/assets/yt-kol-workflow/output/20260721_172846_batch/kol_summary_tables.xlsx"

# 1) tenant token
r = requests.post("https://open.feishu.cn/open-apis/auth/v3/tenant_access_token/internal",
                  json={"app_id": APP_ID, "app_secret": APP_SECRET}, timeout=20)
tok = r.json()["tenant_access_token"]
H = {"Authorization": f"Bearer {tok}"}

# 2) list tables in JB3
r = requests.get(f"https://open.feishu.cn/open-apis/bitable/v1/apps/{JB3}/tables", headers=H, timeout=20)
print("=== JB3 tables ===")
print(json.dumps(r.json(), ensure_ascii=False)[:800])

# 3) fields of default table
r = requests.get(f"https://open.feishu.cn/open-apis/bitable/v1/apps/{JB3}/tables/{TBL}/fields", headers=H, timeout=20)
print("\n=== default table fields ===")
for f in r.json().get("data", {}).get("items", []):
    print(f"  - {f['field_name']!r}  type={f['type']}  id={f['field_id']}")

# 4) Excel sheets & headers
import openpyxl
wb = openpyxl.load_workbook(XLSX, read_only=True, data_only=True)
print("\n=== Excel sheets ===")
for ws in wb.worksheets:
    hdr = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    print(f"  {ws.title}: {ws.max_row-1} rows")
    print("    cols:", hdr)
