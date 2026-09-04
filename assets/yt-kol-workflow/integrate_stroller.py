#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""整合推车 wagon 线：原库 1821 + 新爬 79 -> 过滤汽车/摄像头/足球/旅行EDC -> 生成
独立「网红详情表」(xlsx) 与派生「红人表」(xlsx, 25列)。可审计分类，输出各分类计数。"""
import os, json, datetime
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill

PROJ = os.path.dirname(os.path.abspath(__file__))
LATEST = sorted([f for f in os.listdir(os.path.join(PROJ, "output"))
                 if f.startswith("stroller_crawl_") and f.endswith(".json")])
NEW_PATH = os.path.join(PROJ, "output", LATEST[-1]) if LATEST else None
INF_PATH = os.path.join(PROJ, "local_cache", "influencers.json")

# ---------------- 分类器 ----------------
# 注意：匹配产品字段在当初只匹配 C1 车载摄像头单线，会把 toddler gear / amazon baby products
# 等纯母婴频道也打成 C1，故过滤只用 来源关键词 + 频道初步判断 + 描述 + 频道名，**忽略匹配产品**。
CAR = ["car camera", "autokamera", "auto monitor", "autositz", "car seat",
       "rücksitz", "car essential", "roadtrip", "road trip", "rear facing",
       "auto baby", "汽车", "车载", "baby car", "car monitor", "baby autokamera",
       "new parent car", "car must", "car seat must"]
SOCCER = ["soccer", "足球", "shin guard"]
TRAVEL_EDC = ["airtag", "passport", "business travel", "travel influencer",
              "must-have gadget", "edc"]
BABY = ["baby", "mom", "mother", "mum", "parent", "toddler", "family",
        "stroller", "wagon", "亲子", "母婴", "育儿", "baby gear", "must have",
        "newborn", "infant"]


# 领域黑名单：这些领域的频道即使描述里提了 baby/mom 也非母婴红人，直接排除
EXCLUDE_DOMAIN = ["汽车", "摩托", "健身", "健康", "科技数码", "美食", "烹饪", "户外", "露营"]
# 媒体/娱乐号特征（出现在频道名里基本可判非母婴红人）
MEDIA_NAME = ["music", "studio", "studios", "network", "official", "records",
              "entertainment", "tv", "news", "films", "media", "channel "]


def classify(f):
    def g(k):
        v = f.get(k)
        if isinstance(v, list):
            v = v[0] if v else ""
        return (str(v).lower() if v is not None else "")
    kw, judge, desc, name = g("来源关键词"), g("频道初步判断"), g("频道描述"), g("Channel Name")
    text = f"{kw} {judge} {desc} {name}"  # 不含 匹配产品
    domain = judge.split("领域=")[1].split(";")[0] if "领域=" in judge else ""
    # 育儿词（用于 soccer mom 名判定 / 兜底）
    PARENT_WORD = ("baby", "mom", "mother", "mum", "parent", "toddler", "kid",
                   "family", "亲子", "母婴", "育儿")

    # 1) 汽车/摄像头（来源关键词含 car/auto/kindersitz/monitor/camera 等，或领域汽车/摩托）
    if (any(s in kw for s in ("car", "auto", "kindersitz", "monitor", "camera",
                              "rear facing", "roadtrip", "rücksitz", "autositz"))
            or "汽车" in judge or "摩托" in judge
            or any(s in text for s in CAR)):
        return "car_camera"
    # 2) 媒体/品牌官方号（工作室/影视/VEVO/话题页等，非母婴红人）
    if any(m in name for m in MEDIA_NAME):
        return "other"
    # 3) 足球/足球装备（socor 噪声远大于少数真妈妈号，统一剔除；妈妈群已由新爬+amazon baby/toddler 覆盖）
    if "soccer" in kw or "football" in kw or "shin guard" in kw or "足球" in text:
        return "soccer"
    # 4) 旅行 EDC（AirTag/护照/商务旅行）
    if any(s in kw for s in TRAVEL_EDC) or ("旅行" in judge and "baby" not in text
                                            and "toddler" not in text):
        return "travel_edc"
    # 4) 明确母婴关键词（且未被上面的 car/soccer 拦截）
    if any(w in kw for w in ("baby", "mom", "toddler", "family", "parent",
                             "kid", "amazon baby", "亲子", "母婴", "育儿")):
        return "baby_mom"
    # 5) 兜底：描述含育儿词 且 领域不在黑名单
    if any(w in text for w in BABY) and domain not in EXCLUDE_DOMAIN:
        return "baby_mom"
    return "other"


# ---------------- 载入 ----------------
existing = json.load(open(INF_PATH))["records"]
cat_count = {}
kept_existing = []
for r in existing:
    c = classify(r["fields"])
    cat_count[c] = cat_count.get(c, 0) + 1
    if c == "baby_mom":
        kept_existing.append(r["fields"])

new = json.load(open(NEW_PATH)) if NEW_PATH else []
existing_ids = {r.get("Channel ID") for r in kept_existing}
new_kept = []
new_seen = set()
for r in new:
    cid = r.get("Channel ID")
    if cid in existing_ids or cid in new_seen:
        continue
    new_seen.add(cid)
    # 标注为推车线
    r["匹配产品"] = "WEMOH 2-Seater Stroller Wagon (MCB001-WE)"
    kw = r.get("来源关键词", "")
    r["内容契合类型"] = "评测" if "review" in kw.lower() else ("种草" if "must have" in kw.lower() or "gear" in kw.lower() else "评测/种草")
    r["数据来源"] = "新爬取"
    r["分类"] = "stroller_new"
    new_kept.append(r)

for r in kept_existing:
    r["数据来源"] = "原库"
    r["分类"] = "baby_mom"

all_kept = kept_existing + new_kept
print("=== 原库 1821 分类计数 ===")
for k, v in sorted(cat_count.items(), key=lambda x: -x[1]):
    print(f"  {k:12s} {v}")
print(f"\n原库保留(baby_mom): {len(kept_existing)}")
print(f"新爬保留(去重后):    {len(new_kept)}")
print(f"合并总计:            {len(all_kept)}")

# ---------------- 写 网红详情表 ----------------
DETAIL_COLS = ["Channel ID", "Channel Name", "频道URL", "订阅数", "频道总播放量", "视频总数",
               "国家/地区", "频道描述", "来源关键词", "频道初步判断", "匹配产品", "内容契合类型",
               "品牌匹配度", "亚马逊推广经验", "推广证据", "Amazon Storefront", "代表视频标题",
               "代表视频URL", "代表视频播放量", "代表视频互动率", "联系邮箱", "邮箱来源",
               "开发状态", "开发优先级", "推荐理由", "最新发布日期", "断更评估", "数据来源", "分类"]

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "网红详情表"
ws.append(DETAIL_COLS)
for r in all_kept:
    row = []
    for c in DETAIL_COLS:
        v = r.get(c)
        if isinstance(v, list):
            v = v[0] if v else ""
        row.append(v if v is not None else "")
    ws.append(row)
date = datetime.date.today().strftime("%Y%m%d")
detail_path = os.path.join(PROJ, "output", f"stroller_wagon_网红详情表_{date}.xlsx")
wb.save(detail_path)

# ---------------- 写 红人表 (25列派生) ----------------
RED_COLS = ["Channel ID", "Channel Name", "频道URL", "最新发布日期", "断更评估", "订阅数",
            "国家/地区", "频道初步判断", "联系邮箱", "邮箱来源", "亚马逊推广经验", "Amazon Storefront",
            "推广证据", "匹配产品", "品牌匹配度", "内容契合类型", "开发优先级", "推荐理由",
            "代表视频URL", "代表视频标题", "代表视频互动率", "来源关键词", "开发状态", "开发负责人", "备注"]


def fmt_sub(v):
    try:
        return f"{int(round(float(v))):,}"
    except (TypeError, ValueError):
        return ""


ws2 = wb.create_sheet("红人表")
ws2.append(RED_COLS)
for r in all_kept:
    row = []
    for c in RED_COLS:
        v = r.get(c)
        if isinstance(v, list):
            v = v[0] if v else ""
        if c == "订阅数":
            row.append(fmt_sub(v))
        else:
            row.append(v if v is not None else "")
    ws2.append(row)
red_path = os.path.join(PROJ, "output", f"stroller_wagon_红人表_{date}.xlsx")
wb.save(red_path)

# 简单表头样式
for sh in (ws, ws2):
    for cell in sh[1]:
        cell.font = Font(bold=True)
    sh.freeze_panes = "A2"

print(f"\n✅ 网红详情表 -> {detail_path} ({len(all_kept)} 行)")
print(f"✅ 红人表     -> {red_path} ({len(all_kept)} 行)")
