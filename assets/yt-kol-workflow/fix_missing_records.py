#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Fix script: After write_user_base.py partial failure:
1. Write missing influencer records (200 failed due to network)
2. Restore promo/email fields from local cache for old 497 channels
"""
import os
import openpyxl, json, subprocess, os, sys, time, re

LARK = "/Users/coscod/.workbuddy/binaries/node/versions/22.22.2/bin/lark-cli"
BASE_DIR = "/Users/coscod/WorkBuddy/coscod/YouTube 网红开发工作流系统/assets/yt-kol-workflow"
XLSX = os.path.join(BASE_DIR, "output/summary_german_baby_camera/kol_summary_tables.xlsx")
BASE_TOKEN = os.environ["FEISHU_BASE_TOKEN"]  # 从 .env / 环境变量读取，勿提交真实值
TABLE_INFLUENCERS = "tbl4rnFUM9jJXvCQ"

PROMO_FIELDS = [
    "联系邮箱", "邮箱来源", "邮箱出现视频数", "候选邮箱", "邮箱状态",
    "亚马逊推广经验", "Amazon带货视频数", "推广视频数",
    "Amazon Storefront", "推广证据",
]

SELECT_COLS = {"断更评估", "国家/地区", "邮箱状态", "邮箱来源",
               "亚马逊推广经验", "开发优先级", "开发状态", "来源关键词"}


def _str(v):
    if isinstance(v, list):
        return v[0] if v else ""
    if isinstance(v, (int, float)):
        return v
    return str(v) if v else ""


def run_lark(args, timeout=300):
    env = dict(os.environ)
    env["LARK_CLI_NO_PROXY_WARN"] = "1"
    try:
        p = subprocess.run([LARK] + args, capture_output=True, text=True,
                           timeout=timeout, cwd=BASE_DIR, env=env)
        out = p.stdout.strip()
        err = p.stderr.strip()
        raw = out if out else err
        if raw and not raw.startswith("{"):
            idx = raw.find("\n{")
            if idx >= 0:
                raw = raw[idx+1:].strip()
            else:
                idx = raw.find("{")
                if idx >= 0:
                    raw = raw[idx:]
        try:
            return json.loads(raw) if raw else {}
        except json.JSONDecodeError:
            return {"ok": False, "error": {"message": f"parse: {raw[:200]}"}}
    except subprocess.TimeoutExpired:
        return {"ok": False, "error": {"message": "timeout"}}
    except Exception as e:
        return {"ok": False, "error": {"message": str(e)}}


def get_all_channel_ids():
    """Page through 网红详情表, return {Channel ID: record_id}.
    Uses --offset and --limit for pagination (lark-cli 1.0.73).
    """
    mapping = {}
    offset = 0
    limit = 200
    page = 0
    while True:
        args = ["base", "+record-list", "--as", "user",
                "--base-token", BASE_TOKEN, "--table-id", TABLE_INFLUENCERS,
                "--limit", str(limit), "--offset", str(offset),
                "--format", "json"]
        result = run_lark(args, timeout=120)
        if not result.get("ok"):
            print(f"  [ERROR] record-list: {result.get('error',{}).get('message','')[:200]}")
            break
        d = result.get("data", {})
        rows = d.get("data", [])
        field_names = d.get("fields", [])
        record_ids = d.get("record_id_list", [])

        cid_idx = -1
        for idx, fname in enumerate(field_names):
            if fname == "Channel ID":
                cid_idx = idx
                break

        if cid_idx >= 0:
            for row_idx, row in enumerate(rows):
                if cid_idx < len(row) and row_idx < len(record_ids):
                    cid = _str(row[cid_idx])
                    rid = record_ids[row_idx]
                    if cid and rid:
                        mapping[cid] = rid

        page += 1
        has_more = d.get("has_more", False)
        print(f"  [分页] 第{page}页(offset={offset}): {len(rows)}条, 累计{len(mapping)}, has_more={has_more}", flush=True)
        if not has_more or len(rows) < limit:
            break
        offset += limit
        time.sleep(0.5)
    return mapping


def read_excel_influencers():
    """Read influencers sheet, return (records, headers)."""
    wb = openpyxl.load_workbook(XLSX, read_only=True, data_only=True)
    print(f"  Excel sheets: {wb.sheetnames}", flush=True)
    ws = wb["influencers"]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        wb.close()
        return [], []
    headers = [str(h).strip() if h else "" for h in rows[0]]
    records = []
    for row in rows[1:]:
        if not any(c is not None for c in row):
            continue
        rec = {}
        for i, val in enumerate(row):
            if i < len(headers) and headers[i]:
                rec[headers[i]] = val
        records.append(rec)
    wb.close()
    return records, headers


def build_fields_dict(rec, headers):
    """Convert Excel row to Feishu fields dict."""
    fields = {}
    for h in headers:
        if h not in rec:
            continue
        val = rec[h]
        if val is None:
            continue
        if isinstance(val, float) and val == int(val):
            val = int(val)
        if h in SELECT_COLS:
            val = _str(val) if isinstance(val, list) else str(val)
            if val:
                fields[h] = val
        elif isinstance(val, str):
            val = re.sub(r'<br\s*/?>', '\n', val)
            val = re.sub(r'<[^>]+>', '', val).strip()
            if val:
                fields[h] = val
        elif isinstance(val, (int, float)):
            fields[h] = val
    return fields


def get_feishu_field_names():
    """Get field names from Feishu table."""
    result = run_lark(["base", "+field-list", "--as", "user",
                       "--base-token", BASE_TOKEN, "--table-id", TABLE_INFLUENCERS,
                       "--format", "json"], timeout=120)
    if not result.get("ok"):
        return None
    fields = result.get("data", {}).get("fields", [])
    return [f.get("name", "") for f in fields if f.get("name")]


def write_missing_influencers(existing_ids, excel_records, headers):
    """Write missing records via record-batch-create (matrix format)."""
    missing = [r for r in excel_records
               if str(r.get("Channel ID", "")).strip() and
               str(r.get("Channel ID", "")).strip() not in existing_ids]
    print(f"\n[补写] 缺失网红: {len(missing)} 条")
    if not missing:
        return 0, 0

    # Get Feishu table's actual field names to filter Excel columns
    feishu_fields = get_feishu_field_names()
    if feishu_fields:
        cols = [h for h in headers if h in feishu_fields]
        print(f"  飞书字段: {len(feishu_fields)} 个, Excel匹配: {len(cols)} 个", flush=True)
    else:
        cols = [h for h in headers if h]
        print(f"  无法获取飞书字段, 用Excel全部字段: {len(cols)} 个", flush=True)

    def clean_val(v):
        if v is None:
            return None
        if isinstance(v, float) and v == int(v):
            return int(v)
        if isinstance(v, str):
            v = re.sub(r'<br\s*/?>', '\n', v)
            v = re.sub(r'<[^>]+>', '', v).strip()
            return v if v else None
        # Handle datetime objects from Excel
        import datetime as _dt
        if isinstance(v, (_dt.datetime, _dt.date)):
            return v.strftime("%Y-%m-%d %H:%M:%S")
        return str(v) if v else None

    ok = 0
    fail = 0
    batch_size = 200
    for i in range(0, len(missing), batch_size):
        chunk = missing[i:i+batch_size]
        rows = []
        for rec in chunk:
            row = []
            for col in cols:
                v = rec.get(col)
                if col in SELECT_COLS and isinstance(v, str):
                    v = v.strip() if v.strip() else None
                row.append(clean_val(v))
            rows.append(row)

        payload = {"fields": cols, "rows": rows}
        tmp_path = os.path.join(BASE_DIR, "_fix_batch.json")
        with open(tmp_path, "w", encoding="utf-8") as f:
            json.dump(payload, f, ensure_ascii=False)

        result = run_lark(["base", "+record-batch-create", "--as", "user",
                          "--base-token", BASE_TOKEN, "--table-id", TABLE_INFLUENCERS,
                          "--json", "@_fix_batch.json", "--format", "json"], timeout=300)

        if result.get("ok"):
            ok += len(chunk)
            print(f"  ✅ 批次{i//batch_size+1}: 写入{len(chunk)}条", flush=True)
        else:
            fail += len(chunk)
            print(f"  ✗ 批次{i//batch_size+1}失败: {result.get('error',{}).get('message','')[:200]}", flush=True)
        time.sleep(1)

    return ok, fail


def restore_promo_from_cache(existing_ids):
    """Restore promo/email from local cache (old 497 channels)."""
    cache_path = os.path.join(BASE_DIR, "local_cache", "influencers.json")
    with open(cache_path, "r", encoding="utf-8") as f:
        cache_data = json.load(f)
    cache_records = cache_data["records"]
    print(f"\n[恢复] 缓存: {len(cache_records)}条, 飞书: {len(existing_ids)}条")

    restored = 0
    skipped = 0
    failed = 0

    for i, rec in enumerate(cache_records):
        fields_data = rec.get("fields", {})
        channel_id = _str(fields_data.get("Channel ID", ""))
        if not channel_id:
            skipped += 1
            continue

        record_id = existing_ids.get(channel_id)
        if not record_id:
            skipped += 1
            continue

        # Build update fields from cache
        update_fields = {}
        for field in PROMO_FIELDS:
            if field in fields_data:
                val = _str(fields_data[field])
                if val:
                    update_fields[field] = val

        if not update_fields:
            skipped += 1
            continue

        # record-upsert with --record-id updates existing record
        # --json expects a flat field map (not wrapped in "fields")
        tmp_path = os.path.join(BASE_DIR, "_restore_tmp.json")
        with open(tmp_path, "w", encoding="utf-8") as f:
            json.dump(update_fields, f, ensure_ascii=False)

        result = run_lark(["base", "+record-upsert", "--as", "user",
                          "--base-token", BASE_TOKEN, "--table-id", TABLE_INFLUENCERS,
                          "--record-id", record_id,
                          "--json", "@_restore_tmp.json", "--format", "json"], timeout=300)

        if result.get("ok"):
            restored += 1
        else:
            failed += 1
            if failed <= 3:
                print(f"  [FAIL] {channel_id}: {result.get('error',{}).get('message','')[:200]}", flush=True)

        if (i + 1) % 50 == 0:
            print(f"  进度: {i+1}/{len(cache_records)} (恢复={restored}, 跳过={skipped}, 失败={failed})", flush=True)
        time.sleep(0.3)

    return restored, skipped, failed


def main():
    write_only = "--write-only" in sys.argv
    print("=" * 60)
    print("修复: 补写缺失记录" + ("" if write_only else " + 恢复promo/邮箱"))
    print("=" * 60, flush=True)

    # Step 1: Get existing channel IDs
    print("\n[步骤1] 查询飞书网红详情表...", flush=True)
    existing_ids = get_all_channel_ids()
    print(f"  飞书表现有: {len(existing_ids)} 条", flush=True)

    # Step 2: Read Excel
    print("\n[步骤2] 读取Excel...", flush=True)
    excel_records, headers = read_excel_influencers()
    print(f"  Excel总计: {len(excel_records)} 条, 字段: {len(headers)} 个", flush=True)

    # Step 3: Write missing
    if excel_records:
        ok, fail = write_missing_influencers(existing_ids, excel_records, headers)
        if ok or fail:
            print(f"  补写结果: 成功={ok}, 失败={fail}", flush=True)

    # Step 4: Restore promo/email from cache (skip if --write-only)
    if not write_only:
        print("\n[步骤3] 恢复promo/邮箱...", flush=True)
        restored, skipped, failed = restore_promo_from_cache(existing_ids)
        print(f"  恢复结果: 成功={restored}, 跳过={skipped}, 失败={failed}", flush=True)
    else:
        print("\n[步骤3] 跳过恢复 (--write-only)", flush=True)

    print("\n" + "=" * 60)
    print("修复完成!", flush=True)


if __name__ == "__main__":
    main()
