#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""把本地生成的推车 wagon 两张表（网红详情表 + 红人表）发布到飞书【独立 Base】。
复用 create_us_baby_base.py 的字段推断 + 批量写入逻辑；建一个独立 Base，含两张表。
用法: ./.venv/bin/python publish_stroller_base.py   （不改写已有 VyH0 表）
"""
import os, json, time, re, subprocess, sys
import openpyxl

LARK = "/Users/coscod/.workbuddy/binaries/node/cli-connector-packages/bin/lark-cli"
PROJ = os.path.dirname(os.path.abspath(__file__))
DATE = "20260904"
XLSX_DETAIL = os.path.join(PROJ, "output", f"stroller_wagon_网红详情表_{DATE}.xlsx")
XLSX_RED = os.path.join(PROJ, "output", f"stroller_wagon_红人表_{DATE}.xlsx")
BASE_NAME = "WEMOH推车Wagon网红库"
TBL_DETAIL = "网红详情表"
TBL_RED = "红人表"

SELECT_COLS = {"国家/地区", "开发优先级", "邮箱来源", "亚马逊推广经验", "内容契合类型",
               "匹配产品", "来源关键词", "数据来源", "分类", "开发状态", "断更评估", "邮箱状态"}
URL_HINT = re.compile(r'url|链接', re.I)
DATE_HINT = re.compile(r'日期|时间|date|采集|发布', re.I)
NUM_HINT = re.compile(r'数|量|播放|订阅|互动|率|rate|count|粉丝|views|subs|价格|price|匹配度|总分', re.I)
BAD_OPTS = {"", "?", "？", "未知", "N/A", "na", "none", "null", "-", "(空)"}
BAD_OPTS_LOWER = {o.lower() for o in BAD_OPTS}


def infer_type(col, vals):
    if URL_HINT.search(col):
        return "url"
    if DATE_HINT.search(col) and any(re.search(r'\d{4}[-/]\d{2}[-/]\d{2}', str(v)) for v in vals[:30]):
        return "date"
    if NUM_HINT.search(col):
        num = sum(1 for v in vals if isinstance(v, (int, float)) or
                  (isinstance(v, str) and v.replace('.', '', 1).replace('-', '', 1).lstrip('+').isdigit()))
        if num >= len(vals) * 0.9:
            return "number"
    if col in SELECT_COLS:
        return "select"
    return "text"


def load_sheet(ws):
    rows = list(ws.iter_rows(values_only=True))
    hdr = list(rows[0])
    data = [r for r in rows[1:] if any(c is not None for c in r)]
    col_vals = {c: [r[ci] for r in data if ci < len(r) and r[ci] is not None]
                for ci, c in enumerate(hdr)}
    cols = [c for c in hdr if col_vals.get(c)]
    cols = [c for c in cols if c not in ("多行文本",)]
    specs, records = [], []
    for col in cols:
        vals = col_vals[col]
        t = infer_type(col, vals)
        if t == "select":
            opts = [{"name": str(v)} for v in sorted(
                {str(v).strip() for v in vals if str(v).strip() not in BAD_OPTS})][:100]
            specs.append({"type": "select", "name": col, "options": opts})
        elif t == "url":
            specs.append({"type": "text", "name": col, "style": {"type": "url"}})
        elif t == "date":
            specs.append({"type": "datetime", "name": col, "style": {"format": "yyyy-MM-dd HH:mm"}})
        elif t == "number":
            specs.append({"type": "number", "name": col})
        else:
            specs.append({"type": "text", "name": col})
    for r in data:
        row = {}
        ok = False
        for col in cols:
            ci = hdr.index(col)
            v = r[ci] if ci < len(r) else None
            if v is None or (isinstance(v, str) and v.strip() == ""):
                row[col] = None
                continue
            ok = True
            t = [s for s in specs if s["name"] == col][0]["type"]
            if t == "number":
                try:
                    row[col] = float(v) if ("." in str(v)) else int(v)
                except Exception:
                    row[col] = None
            elif t == "select" and str(v).strip().lower() in BAD_OPTS_LOWER:
                row[col] = None
            else:
                row[col] = str(v)
        if ok:
            records.append(row)
    return cols, specs, records


def run(args, expect_ok=True):
    env = dict(os.environ)
    env["LARK_CLI_NO_PROXY_WARN"] = "1"
    p = subprocess.run([LARK] + args, capture_output=True, text=True, timeout=120, cwd=PROJ, env=env)
    raw = p.stdout.strip() or p.stderr.strip()
    if raw and not raw.startswith("{"):
        i = raw.find("{"); raw = raw[i:] if i >= 0 else raw
    try:
        j = json.loads(raw)
    except Exception:
        print("  ! 非JSON:", raw[:300], p.stderr[:200]); return None
    if expect_ok and not j.get("ok"):
        print("  ! lark-cli 失败:", json.dumps(j, ensure_ascii=False)[:400]); return None
    return j


def batch_write(base_token, table_id, cols, records, pause=0.3):
    ok = fail = 0
    tmp = os.path.join(PROJ, "._stroller_tmp.json")
    for i in range(0, len(records), 200):
        chunk = records[i:i + 200]
        rows = [[r.get(c) for c in cols] for r in chunk]
        body = {"fields": cols, "rows": rows}
        with open(tmp, "w", encoding="utf-8") as f:
            json.dump(body, f, ensure_ascii=False)
        j = run(["base", "+record-batch-create", "--as", "user", "--base-token", base_token,
                 "--table-id", table_id, "--json", "@._stroller_tmp.json", "--format", "json"],
                expect_ok=False)
        if os.path.exists(tmp):
            try: os.remove(tmp)
            except Exception: pass
        if j and j.get("ok"):
            ok += len(chunk)
        else:
            fail += len(chunk)
            print(f"  ✗ 批次失败: {json.dumps(j, ensure_ascii=False)[:300]}")
        if pause: time.sleep(pause)
    return ok, fail


def main():
    print(f"[建 Base] {BASE_NAME}（含『{TBL_DETAIL}』+『{TBL_RED}』）")

    # 表1 网红详情表
    wb1 = openpyxl.load_workbook(XLSX_DETAIL, read_only=True, data_only=True)
    ws1 = wb1[TBL_DETAIL]
    c1, s1, r1 = load_sheet(ws1)
    print(f"  网红详情表: {len(c1)} 字段, {len(r1)} 行")
    # 表2 红人表
    wb2 = openpyxl.load_workbook(XLSX_RED, read_only=True, data_only=True)
    ws2 = wb2[TBL_RED]
    c2, s2, r2 = load_sheet(ws2)
    print(f"  红人表:     {len(c2)} 字段, {len(r2)} 行")

    # 建 Base（首表=网红详情表）
    j = run(["base", "+base-create", "--as", "user", "--name", BASE_NAME,
             "--table-name", TBL_DETAIL, "--fields", json.dumps(s1, ensure_ascii=False),
             "--format", "json"])
    if not j:
        print("建 Base 失败"); sys.exit(1)
    base_token = j["data"]["base"]["base_token"]
    url = j["data"]["base"]["url"]
    print(f"  base_token={base_token}\n  url={url}")

    jt = run(["base", "+table-list", "--as", "user", "--base-token", base_token, "--format", "json"])
    tid_detail = None
    if jt and jt.get("ok"):
        for t in jt["data"].get("tables") or []:
            if t.get("name") == TBL_DETAIL:
                tid_detail = t.get("table_id") or t.get("id")
    print(f"  网红详情表 table_id={tid_detail}")
    ok1, f1 = batch_write(base_token, tid_detail, c1, r1)

    # 建第二表 红人表
    j2 = run(["base", "+table-create", "--as", "user", "--base-token", base_token,
              "--name", TBL_RED, "--fields", json.dumps(s2, ensure_ascii=False), "--format", "json"],
             expect_ok=False)
    tid_red = None
    if j2 and j2.get("ok"):
        # 取 table_id
        def find(d):
            if isinstance(d, dict):
                if "table_id" in d and isinstance(d["table_id"], str): return d["table_id"]
                for v in d.values():
                    r = find(v)
                    if r: return r
            elif isinstance(d, list):
                for v in d:
                    r = find(v)
                    if r: return r
            return None
        tid_red = find(j2.get("data", {}))
    if not tid_red:
        jt2 = run(["base", "+table-list", "--as", "user", "--base-token", base_token, "--format", "json"])
        for t in (jt2["data"].get("tables") or []):
            if t.get("name") == TBL_RED:
                tid_red = t.get("table_id") or t.get("id")
    print(f"  红人表 table_id={tid_red}")
    ok2, f2 = batch_write(base_token, tid_red, c2, r2)

    print(f"\n🎉 完成！Base 链接: https://pcn8zy4grswl.feishu.cn/base/{base_token}")
    print(f"   网红详情表: 写入 {ok1} (失败 {f1})")
    print(f"   红人表:     写入 {ok2} (失败 {f2})")
    print(f"   base_token={base_token}")


if __name__ == "__main__":
    main()
