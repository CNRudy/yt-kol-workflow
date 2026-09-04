#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Re-score 56 newly-found CFC1 influencers with the CFC1 profile, then add them
to the CFC1 Feishu Base (FIhJbkQj / tblpAZFZ5KWywucp).

The fresh batch was accidentally scored against the wrong active profile
(baby car camera). This script re-scores with airtag_passport_holder so the
new 56 rows are consistent with the existing 284.
"""
import os, sys, json
sys.path.insert(0, '.')
os.environ.setdefault('HTTPS_PROXY', 'http://127.0.0.1:7890')
os.environ.setdefault('HTTP_PROXY', 'http://127.0.0.1:7890')
os.environ['LARK_CLI_NO_PROXY_WARN'] = '1'

import openpyxl
from filter import scoring

import write_user_base as wb
wb.LARK = '/Users/coscod/.workbuddy/binaries/node/cli-connector-packages/bin/lark-cli'

BASE_TOKEN = os.environ["FEISHU_BASE_TOKEN_CFC1"]  # 从 .env / 环境变量读取，勿提交真实值
TABLE_ID  = 'tblpAZFZ5KWywucp'
BATCH_DIR = 'output/20260903_164859_batch'

# ---- 1. Load CFC1 profile ----
profiles = json.load(open('product_profiles.json'))
cfc1 = profiles['profiles']['airtag_passport_holder']
print('CFC1 profile name:', cfc1.get('name'))

# ---- 2. Load new influencers + videos ----
wb_inf = openpyxl.load_workbook(f'{BATCH_DIR}/influencers_all.xlsx', read_only=True)
ws = wb_inf.active
rows = list(ws.iter_rows(values_only=True))
H = rows[0]; D = rows[1:]
wb_inf.close()
def ci(name): return H.index(name)

wb_vid = openpyxl.load_workbook(f'{BATCH_DIR}/influencer_videos_all.xlsx', read_only=True)
wsv = wb_vid.active
vrows = list(wsv.iter_rows(values_only=True))
VH = vrows[0]; VD = vrows[1:]
wb_vid.close()
vci = lambda n: VH.index(n)
videos_by_ch = {}
for r in VD:
    ch = str(r[vci('Channel ID')] or '').strip()
    videos_by_ch.setdefault(ch, []).append({
        'title': str(r[vci('Video Title')] or ''),
        'tags': str(r[vci('Tags')] or ''),
        'description': str(r[vci('字幕内容')] or ''),
        'view_count': r[vci('Views')] or 0,
        'engagement_rate': r[vci('互动率(%)')] or 0.0,
    })

# ---- 3. Re-score each new influencer with CFC1 profile ----
records = []
for r in D:
    ch = str(r[ci('Channel ID')] or '').strip()
    if not ch:
        continue
    detail = {
        'amazon_promo_level': str(r[ci('亚马逊推广经验')] or '未发现'),
        'email_status': str(r[ci('邮箱状态')] or ''),
        'rep_video_engagement': float(r[ci('代表视频互动率')] or 0) or 0.0,
        'country': str(r[ci('国家/地区')] or ''),
        'activity_status': '有断更风险' if '断更' in str(r[ci('断更评估')] or '') and '有' in str(r[ci('断更评估')] or '') else '持续更新',
        'channel_description': str(r[ci('频道描述')] or ''),
        'rep_video_title': str(r[ci('代表视频标题')] or ''),
    }
    vids = videos_by_ch.get(ch, [])
    sc = scoring.score_influencer(detail, vids, cfc1)
    rec = {
        'Channel ID': ch,
        '频道名称': str(r[ci('Channel Name')] or ''),
        '频道URL': str(r[ci('频道URL')] or ''),
        '订阅数': r[ci('订阅数')] or 0,
        '国家/地区': str(r[ci('国家/地区')] or '') or None,
        '联系邮箱': str(r[ci('联系邮箱')] or ''),
        '候选邮箱': str(r[ci('候选邮箱')] or ''),
        '邮箱状态': str(r[ci('邮箱状态')] or ''),
        '亚马逊推广经验': str(r[ci('亚马逊推广经验')] or '未发现'),
        'Amazon带货视频数': r[ci('Amazon带货视频数')] or 0,
        'Amazon Storefront': str(r[ci('Amazon Storefront')] or ''),
        '推广证据': str(r[ci('推广证据')] or ''),
        '代表视频URL': str(r[ci('代表视频URL')] or ''),
        '代表视频互动率': float(r[ci('代表视频互动率')] or 0) or 0.0,
        '来源关键词': str(r[ci('来源关键词')] or ''),
        # re-scored match fields
        '匹配产品': sc.get('match_profile', cfc1.get('name')),
        '品牌匹配度': str(sc.get('brand_fit_score', 0)),
        '开发优先级': sc.get('dev_priority', 'C'),
        '推荐理由': sc.get('recommend_reason', ''),
    }
    records.append(rec)

print('准备回填记录数:', len(records))

# ---- 4. Ensure 国家/地区 options cover needed values ----
existing_cids = set(json.load(open('output/cfc1_name_fix/cfc1_existing_cids.json')))
field_list = wb.run(['base','+field-list','--as','user','--base-token',BASE_TOKEN,'--table-id',TABLE_ID,'--format','json'])
country_field = next(f for f in field_list['data']['fields'] if f['name']=='国家/地区')
country_opts = [o['name'] for o in country_field.get('options', [])]
needed = set()
for rec in records:
    c = rec['国家/地区']
    if c and c not in country_opts:
        needed.add(c)
if needed:
    print('需补 国家/地区 选项:', needed)
    existing_full = country_field.get('options', [])
    new_opts = list(existing_full) + [{'name': x} for x in sorted(needed)]
    upd_json = {'name': country_field['name'], 'type': 'select',
                'multiple': country_field.get('multiple', False), 'options': new_opts}
    up = wb.run(['base','+field-update','--as','user','--base-token',BASE_TOKEN,'--table-id',TABLE_ID,
                 '--field-id', country_field['id'], '--format','json','--yes',
                 '--json', json.dumps(upd_json)])
    print('field-update ok=', (up or {}).get('ok'), (up or {}).get('msg',''))

# ---- 5. Build matrix and batch-create ----
fields = ['Channel ID','频道名称','频道URL','订阅数','国家/地区','联系邮箱','候选邮箱','邮箱状态',
          '亚马逊推广经验','Amazon带货视频数','Amazon Storefront','推广证据','代表视频URL',
          '代表视频互动率','来源关键词','匹配产品','品牌匹配度','开发优先级','推荐理由']
matrix_rows = []
for rec in records:
    row = []
    for f in fields:
        v = rec.get(f)
        # select fields must be plain strings; empty select -> '' (skip)
        if f in ('国家/地区','邮箱状态','亚马逊推广经验','来源关键词','匹配产品','开发优先级'):
            v = v if (v and str(v).strip()) else None
        row.append(v)
    matrix_rows.append(row)

created_total = 0
fail = 0
for idx, row in enumerate(matrix_rows):
    j = wb.run(['base','+record-batch-create','--as','user','--base-token',BASE_TOKEN,'--table-id',TABLE_ID,
                '--format','json','--json', json.dumps({'fields': fields, 'rows': [row]})])
    if j and j.get('ok'):
        created_total += 1
    else:
        fail += 1
        rec_ch = records[idx].get('Channel ID')
        # 打出完整报错 + 该记录的 select 值，便于定位
        sel = {f: records[idx].get(f) for f in ('国家/地区','邮箱状态','亚马逊推广经验','来源关键词','匹配产品','开发优先级')}
        print(f'FAIL #{idx} ch={rec_ch} sel={sel}')
        print('   ERR:', json.dumps(j, ensure_ascii=False)[:700] if j else 'None')
        if fail >= 5:
            print('... 已出现 5 个失败，停止以避免重复报错')
            break

print(f'\n=== 完成: 成功 {created_total} / 失败 {fail} (共 {len(matrix_rows)} 条) ===')
